"""
Code will take each cell in a row and assign it to the dictionary variable. Holds information to attach to fill in form code (ref1), clears the dictionary, then loops to the next row.
"""

import openpyxl as xl
### Select what spreadsheet file and reference the sheet name
wb = xl.load_workbook('testlife.xlsx')
sheet = wb['Sheet1']

print(sheet.max_column)
print(sheet.max_row)

client_info = []
up_row = 2

while up_row <= sheet.max_row:
    for row_cells in sheet.iter_rows(min_row= up_row, max_row= up_row):
        for cell in row_cells:
            client_info.append(cell.value)

### Information that will fill out a form
        fill = {
        "lastname": client_info[0],
        "firstName": client_info[1], 
        "DOB": client_info[2],
        
        
    }
### (ref1) develop selenium code below and use dictionary above to fill out any form. 
     ### example: fill["lastname"]       
        print(client_info)
### Dictionary is cleared and then shifts to the following row
        client_info.clear()
        up_row += 1
    