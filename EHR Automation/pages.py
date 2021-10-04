from unittest.case import skip
from selenium.webdriver.support import select, wait
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from elements import BasePageElement
from locators import ClarityClientPrograms, ClarityMainPageLocators, ClarityNewClient 
from selenium.webdriver.common.keys import Keys
import openpyxl as xl
import time
import xlrd

"""Initiation start for the other classes. it all passes through this for the self.driver"""
class BasePage(object):
    """Base class to initialize the base page that will be called from all
    pages"""

    def __init__(self, driver):
        self.driver = driver
        # self.tmp_client_data = {} # This is supposedly used for a temporary placement globally

""" These functions start-up the initial website and the needed input. I call this in the Main --> def setUp """
class StartUp(BasePage):

    def clarity_start(self):
        self.driver = webdriver.Chrome("C:\\Program Files (x86)\\chromedriver.exe")
        self.driver.get("")
        self.driver.maximize_window()
        # This sends keys for login info. Can be mapped to be inputted instead of sitting here
        username = self.driver.find_element(*ClarityMainPageLocators.LOGIN_USER)
        password = self.driver.find_element(*ClarityMainPageLocators.LOGIN_PASS)
        username.send_keys("")
        password.send_keys("")
        # Submit Login
        self.driver.find_element(*ClarityMainPageLocators.LOGIN_GO).click()

"""Functions that manipulate the data for processing"""
class ClarityDataProcess(BasePage):
    """This function 1) reads the excel spreadsheet
    2) converts cells into variables """
    def clarity_new_client(self):
        global fill, client_info, up_row
        wb = xl.load_workbook('Clarity_Test.xlsx')
        sheet = wb['Sheet1']

        """ 2) converts cells into variables """
        client_info = []
        if up_row <= sheet.max_row:
            for row_cells in sheet.iter_rows(min_row= up_row, max_row= up_row):
                for cell in row_cells:
                    client_info.append(cell.value)
                ## Information that will fill out a form
            fill = {
                "id-exist": client_info[5],
                "language": client_info[6],
                "first-name": client_info[7],
                "last-name": client_info[9],
                "middle-name": client_info[8],    
                "q-name" : client_info[10],
                "dob": client_info[18], 
                "q-dob": client_info[19],
                "ssn": client_info[20],
                "q-ssn": client_info[21],
                "suffix/prefix": client_info[11],
                "altname": client_info[12],
                "gender": client_info[13],
                "race-1": client_info[14],
                "race-2": client_info[15],
                "race-3": client_info[16],
                "ethnicity": client_info[17],
                                }
        up_row += 1
        return fill

""" These functions are the general processing for the data.
Global for ContentItems"""
up_row = 2 # <-- This is a global variable for starting the spreadsheet row mapping
class ContentItems(BasePage):   
    """This clicks the menu"""
    def click_menu_button(self):
        element = self.driver.find_element(*ClarityMainPageLocators)
        element.click()

    """Associated with the new client page. Takes the spreadsheet dictionary and filled in the form"""
    def new_client(self):  
        global fill, client_info
        """ 3) Check if they have a HMIS ID, making a new client profile or searching to profile. """
        wait = WebDriverWait(self.driver, 60)   
        if client_info[5] == "Please review":   
            skip
        if client_info[5] == "They don't have a HMIS/Clarity ID":
            wait.until(EC.element_to_be_clickable(ClarityMainPageLocators.GO_CLIENT))
            self.driver.find_element(*ClarityMainPageLocators.GO_CLIENT).click()
            ### This is how the variables are found and inputting into the form
            time.sleep(3)
            self.driver.find_element(*ClarityNewClient.ADD_SSN).send_keys(fill["ssn"])
            Select(self.driver.find_element(*ClarityNewClient.ADD_QSSN)).select_by_visible_text(fill["q-ssn"])
            self.driver.find_element(*ClarityNewClient.ADD_LAST_NAME).send_keys(fill["last-name"])
            self.driver.find_element(*ClarityNewClient.ADD_FIRST_NAME).send_keys(fill["first-name"])
            Select(self.driver.find_element(*ClarityNewClient.ADD_QNAME)).select_by_visible_text(fill["q-name"])
            Select(self.driver.find_element(*ClarityNewClient.ADD_QDOB)).select_by_visible_text(fill["q-dob"])      
            self.driver.find_element(*ClarityNewClient.ADD_DOB).send_keys(fill["dob"])
            self.driver.find_element(*ClarityNewClient.ADD_ALIAS).send_keys(fill["altname"])
            Select(self.driver.find_element(*ClarityNewClient.ADD_GENDER)).select_by_visible_text(fill["gender"])     
            self.driver.find_element(*ClarityNewClient.ADD_MIDDLE_NAME).send_keys(fill["middle-name"])
            Select(self.driver.find_element(*ClarityNewClient.ADD_ETHNICITY)).select_by_visible_text(fill["ethnicity"])
            Select(self.driver.find_element(*ClarityNewClient.ADD_SUFFIX_PREFIX)).select_by_visible_text(fill["suffix/prefix"])
            Select(self.driver.find_element(*ClarityNewClient.ADD_LANGUAGE)).select_by_visible_text(fill["language"])
            self.driver.find_element_by_class_name("placeholder").click()   
            self.driver.find_element_by_xpath("//span[text()='{}']".format(fill["race-1"])).click()
            self.driver.find_element_by_xpath("//span[text()='{}']".format(fill["race-2"])).click()
            self.driver.find_element_by_xpath("//span[text()='{}']".format(fill["race-3"])).click()
            time.sleep(2)
            self.driver.find_element(*ClarityNewClient.ADD_RECORD_CANCEL).click()
            ### REMEMBER NOT TO ACTIVATE THIS DURING TESTING
            # self.driver.find_element(*ClarityNewClient.ADD_RECORD_GO).click()
            
        else:
            wait.until(EC.element_to_be_clickable(ClarityMainPageLocators.SEARCH_CLIENT))
            main_search_bar = self.driver.find_element(*ClarityMainPageLocators.SEARCH_CLIENT)
            main_search_bar.send_keys(fill["id-exist"])
            main_search_bar.send_keys(Keys.ENTER)
        
    def new_program_neighborhood(self):
        
        wait = WebDriverWait(self.driver, 30)
        time.sleep(10)
        program_tab = self.driver.find_element_by_xpath('//a[contains(text(), "Programs")]')
        program_tab.click()
        self.driver.find_element(*ClarityClientPrograms.OUTREACH_NEIGHBORHOOD).click()

        
        # ('//a[contains(text(),{})]'.forma
"""More established codes..? I guess..?"""
class ClarityCompletes(BasePage):
    """This function takes the data, converts the file, pulls the clarity ID or makes a new profile."""
    def clarity_new_client_check(self):
        wb = xl.load_workbook('Clarity_Test.xlsx')
        sheet = wb['Sheet1']
        client_current_row = 2
        client_rows = sheet.max_row
        while client_current_row <= client_rows:
            content_info = ClarityDataProcess.clarity_new_client(self)
            client_stuff = ContentItems.new_client(self)
            content_info
            client_stuff
            client_current_row += 1
            time.sleep(2)
